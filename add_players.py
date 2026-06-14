"""
Dodaje 2 zawodniczki do Ligi 1 kobiet i układa ją na nowo (round-robin 12 osób),
ZACHOWUJĄC kolejkę 1 (te same pary) + nowy mecz Wanda Klimek – Agnieszka Nutowicz.
Pozostałe ligi bez zmian. Wynik nadpisuje Liga_Tenisowa_DANE.xlsx (do re-importu).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import config

PLIK = "Liga_Tenisowa_DANE.xlsx"
LIGA = "Liga 1 kobiet"
NOWE = ["Wanda Klimek", "Agnieszka Nutowicz"]

DATY = {
    1: "15.06-21.06", 2: "22.06-28.06", 3: "29.06-05.07", 4: "06.07-12.07",
    5: "13.07-19.07", 6: "20.07-26.07", 7: "27.07-02.08", 8: "03.08-09.08",
    9: "10.08-16.08", 10: "17.08-23.08", 11: "24.08-30.08",
}


def wczytaj(ws):
    head = [c.value for c in ws[1]]
    out = []
    for r in range(2, ws.max_row + 1):
        d = {head[i]: ws.cell(row=r, column=i + 1).value for i in range(len(head))}
        if str(d.get(head[0]) or "").strip():
            out.append(d)
    return out


wb = openpyxl.load_workbook(PLIK)
mecze = wczytaj(wb["mecze"])
fair = wczytaj(wb["fairplay"])

l1k = [d for d in mecze if str(d.get("Liga")).strip() == LIGA]
inne = [d for d in mecze if str(d.get("Liga")).strip() != LIGA]

# Pary z kolejki 1 (z zachowaniem Gospodarz/Gość) – dokładne nazwiska z danych.
para1 = [(d["Gospodarz"], d["Gość"]) for d in l1k if int(d["Kolejka"]) == 1]
assert len(para1) == 5, f"Oczekiwano 5 par w kolejce 1, jest {len(para1)}"

# Rozsadzenie na kole (circle method): pary kolejki 1 lądują na pozycjach (i, 11-i).
P = [None] * 12
for i, (g, go) in enumerate(para1):
    P[i] = g
    P[11 - i] = go
P[5] = NOWE[0]
P[6] = NOWE[1]
assert all(P), "Brak obsadzenia wszystkich pozycji"
assert len(set(P)) == 12, "Powtórzone nazwisko – sprawdź dane"

# Generacja 11 kolejek round-robin.
N = 12
arr = list(range(N))
nowe_l1k = []
for rnd in range(N - 1):
    kol = rnd + 1
    for i in range(N // 2):
        nowe_l1k.append({
            "Liga": LIGA, "Kolejka": kol, "Data": DATY[kol],
            "Gospodarz": P[arr[i]], "Gość": P[arr[N - 1 - i]],
        })
    arr = [arr[0]] + [arr[-1]] + arr[1:-1]

# --- Weryfikacja ---
r1_new = {frozenset((m["Gospodarz"], m["Gość"])) for m in nowe_l1k if m["Kolejka"] == 1}
r1_old = {frozenset(p) for p in para1} | {frozenset(NOWE)}
assert r1_new == r1_old, "Kolejka 1 się zmieniła!"
# każdy gra 11 meczów, raz na kolejkę
from collections import Counter
licz = Counter()
for m in nowe_l1k:
    licz[m["Gospodarz"]] += 1
    licz[m["Gość"]] += 1
assert all(v == 11 for v in licz.values()), f"Ktoś nie gra 11 meczów: {licz}"
for kol in range(1, 12):
    osoby = [m["Gospodarz"] for m in nowe_l1k if m["Kolejka"] == kol] + \
            [m["Gość"] for m in nowe_l1k if m["Kolejka"] == kol]
    assert len(osoby) == len(set(osoby)) == 12, f"Kolejka {kol} ma duplikaty/braki"

# --- Złożenie wszystkich meczów (ligi w kolejności z config, sort po kolejce) ---
po_lidze = {LIGA: nowe_l1k}
for d in inne:
    po_lidze.setdefault(str(d["Liga"]).strip(), []).append(d)

final = []
for liga in config.LIGI.values():
    grupa = sorted(po_lidze.get(liga, []), key=lambda d: int(d["Kolejka"]))
    final += grupa
for i, d in enumerate(final, 1):
    d["ID"] = i

# Fair Play: dotychczasowi + 2 nowe (w grupie Ligi 1 kobiet), reszta bez zmian.
fp_new = []
dodano = False
for row in fair:
    fp_new.append(row)
for n in NOWE:
    fp_new.append({"Zawodnik": n, "Liga": LIGA, "Punkty Fair Play": 0})

# --- Zapis nowego pliku ---
out = openpyxl.Workbook()


def naglowki(ws, kolumny):
    fill = PatternFill("solid", fgColor="7CA87C")
    for i, k in enumerate(kolumny, 1):
        c = ws.cell(row=1, column=i, value=k)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(k) + 2)
    ws.freeze_panes = "A2"


ws_m = out.active
ws_m.title = config.WS_MECZE
naglowki(ws_m, config.KOL_MECZE)
for r, d in enumerate(final, 2):
    for c, k in enumerate(config.KOL_MECZE, 1):
        ws_m.cell(row=r, column=c, value=d.get(k, ""))
for kol in ("Set1", "Set2", "SuperTB", "Wynik"):
    c = config.KOL_MECZE.index(kol) + 1
    for r in range(2, len(final) + 2):
        ws_m.cell(row=r, column=c).number_format = "@"

ws_z = out.create_sheet(config.WS_ZGLOSZENIA)
naglowki(ws_z, config.KOL_ZGLOSZENIA)

ws_f = out.create_sheet(config.WS_FAIRPLAY)
naglowki(ws_f, config.KOL_FAIRPLAY)
for r, d in enumerate(fp_new, 2):
    for c, k in enumerate(config.KOL_FAIRPLAY, 1):
        ws_f.cell(row=r, column=c, value=d.get(k, ""))

out.save(PLIK)

safe = lambda s: str(s).encode("ascii", "replace").decode()
print("OK. Mecze razem:", len(final), "| Liga 1 kobiet:", len(nowe_l1k),
      "| Fair Play:", len(fp_new))
print("Kolejka 1 (Liga 1 kobiet):")
for m in nowe_l1k:
    if m["Kolejka"] == 1:
        print("  ", safe(m["Gospodarz"]), "vs", safe(m["Gość"]))
