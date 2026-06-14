"""
Konwersja 'Liga uczestnicy.xlsx' -> 'Liga_Tenisowa_DANE.xlsx' (gotowe do importu
do arkusza Google: zakładki mecze / zgloszenia / fairplay).

- nazwiska ujednolicone do formatu "Imię Nazwisko",
- poprawne nazwy lig, unikalne ID, daty kolejek jako zakres tygodnia,
- Fair Play: wszyscy zawodnicy z 0 pkt na start (ustawiasz ręcznie).
"""

import re
import unicodedata

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import config

ZRODLO = "Liga uczestnicy.xlsx"
WYNIK = "Liga_Tenisowa_DANE.xlsx"

# zakładka rosteru -> (kanoniczna nazwa ligi, zakładka terminarza)
LIGI = {
    "Liga 1 M": ("Liga 1 mężczyzn", "Terminarz Liga 1 M"),
    "Liga 2 M": ("Liga 2 mężczyzn", "Terminarz Liga 2 M"),
    "Liga 3 M": ("Liga 3 mężczyzn", "Terminarz Liga 3 M"),
    "Liga 1 K": ("Liga 1 kobiet", "Terminarz 1 Liga K"),
    "Liga 2 K": ("Liga 2 kobiet", "Terminarz 2 Liga K"),
    "Liga 3 K": ("Liga 3 kobiet", "Terminarz 3 Liga K"),
}

safe = lambda s: str(s).encode("ascii", "replace").decode()


def s(v):
    return "" if v is None else str(v).strip()


def tc(x):  # title-case z obsługą myślników i polskich znaków
    return " ".join(
        "-".join(seg[:1].upper() + seg[1:].lower() for seg in w.split("-"))
        for w in x.split()
    )


def norm(x):
    x = unicodedata.normalize("NFKD", str(x)).encode("ascii", "ignore").decode().lower()
    return re.sub(r"\s+", " ", x).strip()


def main():
    wb = openpyxl.load_workbook(ZRODLO, data_only=True)

    # 1) Mapy nazwisk per liga + lista zawodników
    mapy, zawodnicy = {}, {}
    for tab, (kanon, _) in LIGI.items():
        ws = wb[tab]
        m, lista = {}, []
        for row in ws.iter_rows(min_row=2, values_only=True):
            naz, imie = s(row[0]), s(row[1])
            if not naz:
                continue
            c = f"{tc(imie)} {tc(naz)}".strip()
            m[norm(f"{naz} {imie}")] = c
            m[norm(f"{imie} {naz}")] = c
            lista.append(c)
        mapy[kanon] = m
        zawodnicy[kanon] = lista

    # 2) Terminarze -> mecze
    def odczytaj(tab, kanon):
        ws = wb[tab]
        m = mapy[kanon]
        out, kol, data, niez = [], None, None, []
        for row in ws.iter_rows(values_only=True):
            a = s(row[0])
            b = s(row[1]) if len(row) > 1 else ""
            if a.startswith("Bilans") or a in ("Zawodnik", "Zawodniczka"):
                break  # koniec terminarza, dalej tabela bilansu
            mk = re.match(r"Kolejka\s*(\d+)", a)
            if mk:
                kol = int(mk.group(1))
                reszta = a[mk.end():].lstrip(" –—-:").strip()
                data = reszta.replace("–", "-").replace("—", "-")
                continue
            low = a.lower()
            if not a or low.startswith("pauza") or low == "gospodarz" or low.startswith("terminarz"):
                continue
            if a and b:
                g = m.get(norm(a))
                go = m.get(norm(b))
                if g is None:
                    niez.append(a)
                if go is None:
                    niez.append(b)
                out.append((kol, data, g or tc(a), go or tc(b)))
        return out, niez

    wyniki, wszystkie_niez = {}, []
    for tab, (kanon, term) in LIGI.items():
        mecze, niez = odczytaj(term, kanon)
        wyniki[kanon] = mecze
        wszystkie_niez += [(kanon, n) for n in niez]

    # 3) Budowa pliku wynikowego
    out = openpyxl.Workbook()

    def naglowki(ws, kolumny):
        fill = PatternFill("solid", fgColor="7CA87C")
        for i, k in enumerate(kolumny, 1):
            c = ws.cell(row=1, column=i, value=k)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = fill
            ws.column_dimensions[get_column_letter(i)].width = max(12, len(k) + 2)
        ws.freeze_panes = "A2"

    ws_m = out.active
    ws_m.title = config.WS_MECZE
    naglowki(ws_m, config.KOL_MECZE)
    idx_map = {k: i + 1 for i, k in enumerate(config.KOL_MECZE)}
    r = 2
    licznik = 1
    for kanon in config.LIGI.values():  # zachowaj kolejność lig z config
        for kol, data, g, go in wyniki.get(kanon, []):
            wartosci = {"ID": licznik, "Liga": kanon, "Kolejka": kol,
                        "Data": data, "Gospodarz": g, "Gość": go}
            for k, col in idx_map.items():
                ws_m.cell(row=r, column=col, value=wartosci.get(k, ""))
            licznik += 1
            r += 1
    ostatni = r - 1
    for kol in ("Set1", "Set2", "SuperTB", "Wynik"):
        col = idx_map[kol]
        for row in range(2, ostatni + 1):
            ws_m.cell(row=row, column=col).number_format = "@"

    ws_z = out.create_sheet(config.WS_ZGLOSZENIA)
    naglowki(ws_z, config.KOL_ZGLOSZENIA)

    ws_f = out.create_sheet(config.WS_FAIRPLAY)
    naglowki(ws_f, config.KOL_FAIRPLAY)
    rf = 2
    for kanon in config.LIGI.values():
        for zaw in zawodnicy.get(kanon, []):
            ws_f.cell(row=rf, column=1, value=zaw)
            ws_f.cell(row=rf, column=2, value=kanon)
            ws_f.cell(row=rf, column=3, value=0)
            rf += 1

    out.save(WYNIK)

    # 4) Raport
    print("LIGA                 zaw  mecz")
    for kanon in config.LIGI.values():
        print(f"  {safe(kanon):18} {len(zawodnicy.get(kanon,[])):3}  {len(wyniki.get(kanon,[])):4}")
    print("RAZEM meczow:", licznik - 1, "| zawodnikow:", rf - 2)
    print("\nNIEZMAPOWANE (ma byc puste):")
    if wszystkie_niez:
        for liga, n in wszystkie_niez:
            print("  ", safe(liga), "->", safe(n))
    else:
        print("  (brak - wszystko zmapowane)")
    print(f"\nZapisano: {WYNIK}")


if __name__ == "__main__":
    main()
